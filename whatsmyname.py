import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import csv

def check_site(site, username, headers):
    site_name = site["name"]
    uri_check = site["uri_check"].format(account=username)
    try:
        res = requests.get(uri_check, headers=headers, timeout=10)
        estring_pos = site["e_string"] in res.text
        estring_neg = site["m_string"] in res.text

        if res.status_code == site["e_code"] and estring_pos and not estring_neg:
            return site_name, uri_check
    except:
        pass
    return None

def generate_html_report(username, found_sites):
    html_content = f"""
    <html>
    <head>
        <title>WhatsMyName Report for {username}</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }}
            th {{
                background-color: #f2f2f2;
            }}
        </style>
    </head>
    <body>
        <h1>WhatsMyName Report for {username}</h1>
        <table>
            <tr>
                <th>Website Name</th>
                <th>Profile URL</th>
            </tr>"""
    for site in found_sites:
        html_content += f"""
            <tr>
                <td>{site['name']}</td>
                <td><a href="{site['url']}" target="_blank">{site['url']}</a></td>
            </tr>"""
    html_content += """
        </table>
    </body>
    </html>"""

    filename = f"whatsmyname_report_{username}.html"
    with open(f"static/{filename}", "w") as report_file:
        report_file.write(html_content)
    
    return filename

def generate_excel_report(username, found_sites):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"WhatsMyName Report - {username}"

    ws['A1'] = "Website Name"
    ws['B1'] = "Profile URL"

    for row, site in enumerate(found_sites, start=2):
        ws[f'A{row}'] = site['name']
        ws[f'B{row}'] = site['url']

    filename = f"whatsmyname_report_{username}.xlsx"
    wb.save(f"static/{filename}")
    return filename

def generate_pdf_report(username, found_sites):
    filename = f"whatsmyname_report_{username}.pdf"
    doc = SimpleDocTemplate(f"static/{filename}", pagesize=letter)
    
    data = [["Website Name", "Profile URL"]]
    for site in found_sites:
        data.append([site['name'], site['url']])

    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)
    
    elements = []
    elements.append(table)
    
    doc.build(elements)
    return filename

def generate_csv_report(username, found_sites):
    filename = f"whatsmyname_report_{username}.csv"
    with open(f"static/{filename}", 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)
        csvwriter.writerow(["Website Name", "Profile URL"])
        for site in found_sites:
            csvwriter.writerow([site['name'], site['url']])
    return filename
